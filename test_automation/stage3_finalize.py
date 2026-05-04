import openpyxl
import json

# Load the rationale data we saved in Stage 1
with open('rationale_data.json', 'r', encoding='utf-8') as f:
    rationale_data = json.load(f)

# Build a lookup dict by TC ID
rationale_map = { row['ID']: (row['Cat'], row['Rationale']) for row in rationale_data }

file_path = "Assignment 1 - Test cases.xlsx"

wb = openpyxl.load_workbook(file_path)

# Find the correct sheet
if ' Test cases' in wb.sheetnames:
    ws = wb[' Test cases']
elif 'Test cases' in wb.sheetnames:
    ws = wb['Test cases']
else:
    ws = wb.active

print(f"Working on sheet: '{ws.title}'")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Read the header row to find column positions
header_row_idx = None
header_map = {}
for row in ws.iter_rows(min_row=1, max_row=5):
    for cell in row:
        if cell.value and str(cell.value).strip() in ['TC ID', 'Input length type', 'Input', 'Expected output', 'Actual output', 'Status']:
            header_row_idx = cell.row
            break
    if header_row_idx:
        break

if not header_row_idx:
    header_row_idx = 1

print(f"Header row: {header_row_idx}")

# Read all headers
for cell in ws[header_row_idx]:
    if cell.value:
        header_map[str(cell.value).strip()] = cell.column

print(f"Headers found: {list(header_map.keys())}")

# Determine where to put the new columns
# Find the last used column
last_col = ws.max_column

# Check if 'Singlish input types covered' already exists
singlish_col = None
rationale_col = None
for col_idx in range(1, last_col + 1):
    val = ws.cell(row=header_row_idx, column=col_idx).value
    if val and 'Singlish input types covered' in str(val):
        singlish_col = col_idx
    if val and 'Evidence or rationale' in str(val):
        rationale_col = col_idx

if not singlish_col:
    singlish_col = last_col + 1
    ws.cell(row=header_row_idx, column=singlish_col).value = 'Singlish input types covered'

if not rationale_col:
    rationale_col = singlish_col + 1
    ws.cell(row=header_row_idx, column=rationale_col).value = 'Evidence or rationale for the input type covered'

print(f"'Singlish input types covered' column: {singlish_col}")
print(f"'Evidence or rationale' column: {rationale_col}")

# Find TC ID column
tc_id_col = header_map.get('TC ID', 1)

# Fill the two columns for each data row
filled = 0
for row_idx in range(header_row_idx + 1, ws.max_row + 1):
    tc_id_cell = ws.cell(row=row_idx, column=tc_id_col)
    tc_id = str(tc_id_cell.value).strip() if tc_id_cell.value else ''
    if not tc_id or tc_id == 'None':
        continue

    if tc_id in rationale_map:
        cat, rationale = rationale_map[tc_id]
        ws.cell(row=row_idx, column=singlish_col).value = cat
        ws.cell(row=row_idx, column=rationale_col).value = "Rationale: " + rationale
        filled += 1
        print(f"  Row {row_idx} ({tc_id}): {cat}")
    else:
        print(f"  WARNING: No rationale found for TC ID '{tc_id}'")

wb.save(file_path)
print(f"\nStage 3 complete! Filled {filled} rows. Saved to: {file_path}")
