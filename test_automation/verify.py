import openpyxl

file_path = "IT23666610_Assignment 1 - Test cases.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("=== Excel File Verification ===")
print(f"Sheet: '{ws.title}'")
print(f"Total rows (including header): {ws.max_row}")
print(f"Total columns: {ws.max_column}")
print()

# Print headers
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print("Headers:", headers)
print()

# Sample first 3 rows
print("=== First 3 data rows ===")
for row_idx in range(2, 5):
    row_data = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Row {row_idx}: {row_data}")

# Count statuses
print()
statuses = {}
for row_idx in range(2, ws.max_row + 1):
    status = ws.cell(row=row_idx, column=6).value  # Status column
    statuses[status] = statuses.get(status, 0) + 1

print("=== Status Summary ===")
for status, count in statuses.items():
    print(f"  {status}: {count}")

# Check that columns 7 and 8 are filled
empty_cat = 0
empty_rat = 0
for row_idx in range(2, ws.max_row + 1):
    if not ws.cell(row=row_idx, column=7).value:
        empty_cat += 1
    if not ws.cell(row=row_idx, column=8).value:
        empty_rat += 1

print(f"\nRows missing 'Singlish input types covered': {empty_cat}")
print(f"Rows missing 'Evidence or rationale': {empty_rat}")
print("\n✅ Verification complete!")
